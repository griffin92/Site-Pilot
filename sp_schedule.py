"""Timeline generation and Gantt template fill.

DATE MATH LIVES IN PYTHON, NOT THE PROMPT. Language models are unreliable at
business-day arithmetic -- they'll cheerfully put a task start on a Saturday
or drift a chain of dependencies by days. The AI proposes task names,
durations, and dependencies; Python computes every actual date.
"""
import io
from datetime import timedelta

import openpyxl

import sp_ai as ai
import sp_prompts as prompts
from sp_config import (GANTT_TEMPLATE_PATH, GANTT_TASK_START_ROW,
                    GANTT_MAX_TASK_ROWS, DEEP_SCAN_BATCH)


# ------------------------------------------------------------- date helpers

def add_business_days(start, work_days):
    """Mirrors the template's WORKDAY(start, work_days-1)."""
    if work_days <= 1:
        return start
    d, remaining = start, work_days - 1
    while remaining > 0:
        d += timedelta(days=1)
        if d.weekday() < 5:
            remaining -= 1
    return d


def next_business_day(d):
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def schedule_tasks(project_start, tasks):
    end_by_wbs = {}
    for t in tasks:
        preds = t.get("predecessors") or []
        pred_ends = [end_by_wbs[p] for p in preds if p in end_by_wbs]
        start = next_business_day(max(pred_ends) + timedelta(days=1)) if pred_ends else project_start
        end = add_business_days(start, max(1, int(t.get("work_days", 1))))
        t["start"], t["end"] = start, end
        end_by_wbs[t["wbs"]] = end
    return tasks


# ---------------------------------------------------------------- generation

def generate_timeline(file_bytes, pages, start_date, target_weeks=None, label_fn=None):
    """target_weeks: optional deadline. The model sequences toward it, but is
    told to flag infeasibility rather than silently shrink task durations --
    a schedule that only fits because the AI shaved every duration is worse
    than no schedule at all."""
    user = (f"Analyze these drawings. Project start date is {start_date.strftime('%b %d, %Y')}. "
            f"Produce a chronological construction timeline broken into discrete sequential "
            f"tasks, grouped by phase, with a working-day duration and prerequisites for each.")

    if target_weeks:
        work_days = int(target_weeks) * 5
        user += (
            f"\n\nTARGET DURATION: the project must reach substantial completion within "
            f"{int(target_weeks)} weeks ({work_days} working days) of the start date.\n"
            f"Sequence and overlap trades to hit this target where it is safe and realistic "
            f"to do so -- run independent trades in parallel, and note where crew size or "
            f"shift work is what makes the date achievable.\n"
            f"CRITICAL: do NOT shorten individual task durations below what the work actually "
            f"takes just to make the arithmetic fit. If this scope cannot responsibly be built "
            f"in {int(target_weeks)} weeks, begin your response with a section headed "
            f"'SCHEDULE FEASIBILITY' stating that plainly, give your realistic minimum "
            f"duration in weeks, and name the specific constraints driving it (long-lead "
            f"items, inspection holds, cure times, trade dependencies). Then give the best "
            f"achievable schedule."
        )

    return ai.batched_scan(
        file_bytes, pages, prompts.SCHEDULER, user,
        batch_size=DEEP_SCAN_BATCH, label_fn=label_fn,
        progress_label="Building timeline",
    )


def duration_summary(project_start, tasks):
    """Actual duration of the scheduled tasks, computed in Python.

    This is the honest check on the target: the AI proposes durations and
    dependencies, but only the real date arithmetic tells you whether the
    plan lands inside the deadline.
    """
    if not tasks:
        return None
    scheduled = schedule_tasks(project_start, [dict(t) for t in tasks])
    finish = max(t["end"] for t in scheduled)
    calendar_days = (finish - project_start).days + 1
    weeks = calendar_days / 7.0
    work_days = sum(max(1, int(t.get("work_days", 1))) for t in scheduled)
    return {
        "finish": finish,
        "calendar_days": calendar_days,
        "weeks": weeks,
        "task_count": len(scheduled),
        "total_work_days": work_days,
    }


def extract_tasks(timeline_text):
    """Free text -> structured task list via JSON mode."""
    user = f"""Extract every task from this construction timeline into a JSON array.
Each item must have exactly:
- "wbs": sequential integer starting at 1
- "task": short task name
- "work_days": duration in working days (integer, min 1)
- "predecessors": array of WBS integers it depends on (empty if none)

Output ONLY the raw JSON array.

Timeline:
{timeline_text}"""
    return ai.generate_json(
        [user],
        "You are a data extraction engine converting construction schedules into structured task lists.",
        default=[],
    )


def fill_gantt(project_start, tasks):
    """Writes only the plain input cells of the Vertex42 template; every
    locked formula (End, Cal Days, Days Done/Left) is left untouched so the
    workbook behaves exactly as the original does.

    Returns (xlsx_bytes, truncated_flag).
    """
    wb = openpyxl.load_workbook(GANTT_TEMPLATE_PATH)
    ws = wb["GanttChart"]
    ws["G6"] = project_start

    scheduled = schedule_tasks(project_start, tasks)
    truncated = len(scheduled) > GANTT_MAX_TASK_ROWS
    scheduled = scheduled[:GANTT_MAX_TASK_ROWS]

    for i in range(GANTT_MAX_TASK_ROWS):
        row = GANTT_TASK_START_ROW + i
        if i < len(scheduled):
            t = scheduled[i]
            ws.cell(row=row, column=1).value = t["wbs"]
            ws.cell(row=row, column=2).value = t["task"]
            ws.cell(row=row, column=4).value = ",".join(str(p) for p in t.get("predecessors", [])) or None
            ws.cell(row=row, column=7).value = t["start"]
            ws.cell(row=row, column=9).value = max(1, int(t.get("work_days", 1)))
            ws.cell(row=row, column=10).value = 0
        else:
            # openpyxl treats .cell(value=None) as a no-op, so clear via .value
            for col in (1, 2, 4, 7, 9, 10):
                ws.cell(row=row, column=col).value = None

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), truncated
